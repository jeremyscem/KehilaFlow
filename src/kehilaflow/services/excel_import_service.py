from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from kehilaflow.api.schemas.imports import (
    ExcelPreviewResponse,
    ExcelValue,
)


def _normalize_value(value: object) -> ExcelValue:
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def _get_columns(
    first_row: tuple[object, ...],
) -> list[tuple[int, str]]:
    """
    Return only columns that have a real title.

    Example:
    DATE | NOM | DU | [empty] | [empty]

    becomes:
    [(0, "DATE"), (1, "NOM"), (2, "DU")]

    Columns without a title are completely ignored.
    """
    columns: list[tuple[int, str]] = []

    for index, value in enumerate(first_row):
        if value is None:
            continue

        column_name = str(value).strip()

        if not column_name:
            continue

        columns.append(
            (
                index,
                column_name,
            )
        )

    return columns


def _build_row(
    row: tuple[object, ...],
    columns: list[tuple[int, str]],
) -> dict[str, ExcelValue]:
    normalized_row: dict[str, ExcelValue] = {}

    for index, column in columns:
        value = row[index] if index < len(row) else None

        normalized_value = _normalize_value(value)

        # Empty financial cells are considered zero.
        if column.upper() in {"DU", "PAYE"} and normalized_value is None:
            normalized_value = 0

        normalized_row[column] = normalized_value

    return normalized_row


def preview_excel(
    file_bytes: bytes,
    file_name: str,
    limit: int = 10,
) -> ExcelPreviewResponse:
    workbook = load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active
    rows_iterator = sheet.iter_rows(values_only=True)

    first_row = next(rows_iterator, None)

    if first_row is None:
        workbook.close()

        return ExcelPreviewResponse(
            file_name=file_name,
            sheet_name=sheet.title,
            columns=[],
            rows=[],
            total_rows=0,
        )

    columns_with_indexes = _get_columns(first_row)

    columns = [column for _, column in columns_with_indexes]

    preview_rows: list[dict[str, ExcelValue]] = []

    total_rows = 0

    for row in rows_iterator:
        normalized_row = _build_row(
            row=row,
            columns=columns_with_indexes,
        )

        # Ignore rows that contain no data
        # in any titled column.
        if not any(value is not None for value in normalized_row.values()):
            continue

        total_rows += 1

        if len(preview_rows) < limit:
            preview_rows.append(normalized_row)

    sheet_name = sheet.title

    workbook.close()

    return ExcelPreviewResponse(
        file_name=file_name,
        sheet_name=sheet_name,
        columns=columns,
        rows=preview_rows,
        total_rows=total_rows,
    )


def read_excel_rows(
    file_bytes: bytes,
) -> tuple[
    list[str],
    list[dict[str, ExcelValue]],
]:
    workbook = load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active
    rows_iterator = sheet.iter_rows(values_only=True)

    first_row = next(rows_iterator, None)

    if first_row is None:
        workbook.close()
        return [], []

    columns_with_indexes = _get_columns(first_row)

    columns = [column for _, column in columns_with_indexes]

    rows: list[dict[str, ExcelValue]] = []

    for row in rows_iterator:
        normalized_row = _build_row(
            row=row,
            columns=columns_with_indexes,
        )

        # Ignore rows that contain no data
        # in titled columns.
        if not any(value is not None for value in normalized_row.values()):
            continue

        rows.append(normalized_row)

    workbook.close()

    return columns, rows
